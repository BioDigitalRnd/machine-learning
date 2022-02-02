import openpyxl as xl

filename = "bidmc_numerics.xlsm"

wb = xl.load_workbook(filename)


for x in range(1, 54):
    sheet_name = f'bidmc_{str(x).zfill(2)}_Numerics'
    sheet = wb[sheet_name]

    cell_age_text = sheet.cell(2, 23)
    cell_age_value = sheet.cell(2, 24)
    cell_gender_text = sheet.cell(3, 23)
    cell_gender_value = sheet.cell(3, 24)
    cell_patient_no_text = sheet.cell(5, 23)
    cell_patient_no_value = sheet.cell(5, 24)

    cell_age_text.value = "Age"
    cell_gender_text.value = "Gender"
    cell_patient_no_text.value = "Patient"
    cell_patient_no_value.value = x
wb.save(filename)



