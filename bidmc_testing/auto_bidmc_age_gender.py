# download openpyxl package from pypi.org

import openpyxl as xl
#
# #bidmc_numerics.xlsm
#
filename = "bidmc_numerics.xlsx"
#
# # sheet_name = "bidmc_XX_Numerics"
# # sheet = wb[sheet_name]
# # sheets = []
#
wb = xl.load_workbook(filename)


for x in range(1, 54):
    sheet_name = f'bidmc_{str(x).zfill(2)}_Numerics'
    sheet = wb[sheet_name]

    cell_age_text = sheet.cell(2, 23)
    cell_age_value = sheet.cell(2, 24)
    cell_gender_text = sheet.cell(3, 23)
    cell_gender_value = sheet.cell(3, 24)
    cell_patient_no_text = sheet.cell(5, 23)
    cell_patient_no_value = sheet.cell(5, 24)
    cell_patient_no_value.value = x
    cell_age_text.value = "Age"
    cell_gender_text.value = "Gender"
    cell_patient_no_text.value = "Patient"

    with open(f'bidmc_{str(x).zfill(2)}_Fix.txt') as pt:
        for rec in pt:
            if rec[0:3] == 'Age':
                cell_age_value.value = rec[5:7]
            if rec[0:3] == 'Gen':
                if rec[8] == 'M':
                    cell_gender_value.value = rec[8]
                elif rec[8] == 'F':
                    cell_gender_value.value = rec[8]
                else:
                    cell_gender_value.value = "NaN"

wb.save(filename)




# print(patient_age['patient18'])
